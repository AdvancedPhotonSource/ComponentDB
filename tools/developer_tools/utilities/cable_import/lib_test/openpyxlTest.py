from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws['A1'] = "Name"
ws['B1'] = "Description"
ws['C1'] = "Contact Info"
ws['D1'] = "URL"
ws['A2'] = "CommScope"
wb.save("/home/craig/projects/20200407-preimport-scripts/openpyxl-test.xlsx")
