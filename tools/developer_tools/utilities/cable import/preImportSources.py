#
# Prepares spreadsheet of Sources for import to CDB.
#
# Inputs:
#   * xlsx file with cleaned up cable specs tab from cable data collection workbook
#       (delete all but single team's non-empty rows)
#
# Outputs:
#   * official CDB Sources import format xlsx file
#
# Processing:
#   * collect the unique set of manufacturers from the input spreadsheet (there are many duplicates)
#   * iterate through the unique set of manufacturers
#       - check if there is an existing cdb source for the manufacturer and if so find its id
#       - check to see if there is a close match using find - I have a hunch there are variations with "Inc.", "Co." etc
#       - if the source does not exist in CDB, add a row for it to the output spreadsheet
#           - the CDB Sources import mechanism will ignore duplicates, if we want to take advantage of that
#       - log a message about the Source, if it exists already, if there are existing variations on the name, etc

import argparse
import sys
import openpyxl
import xlsxwriter

from CdbApiFactory import CdbApiFactory
from cdbApi import ApiException


def main():

    # process command line

    parser = argparse.ArgumentParser()
    parser.add_argument("inputFile", help="Data collection workbook xlsx file with 'cable specs' tab")
    parser.add_argument("outputFile", help="Official CDB Sources import format xlsx file")
    parser.add_argument("--logFile", help="File for log output", required=True)
    args = parser.parse_args()
    print("using inputFile: %s" % args.inputFile)
    print("using outputFile: %s" % args.outputFile)
    print("using logFile: %s" % args.logFile)

    # process input spreadsheet

    input_book = openpyxl.load_workbook(filename=args.inputFile, read_only=True, data_only=True)
    input_sheet = input_book.active
    input_sheet.reset_dimensions()
    print("input spreadsheet dimensions: %s" % input_sheet.calculate_dimension(force=True))

    if input_sheet.max_row < 2:
        sys.exit("no data in inputFile: %s" % args.inputFile)
    if input_sheet.max_column != 16:
        sys.exit("inputFile %s doesn't contain expected number of columns" % args.inputFile)

    row_count = 0
    manufacturers = set()
    for row in input_sheet.rows:
        row_count = row_count + 1
        if row_count == 1:
            continue
        manufacturer = row[4].value
        print("row %d found manufacturer: %s" % (row_count, manufacturer))
        manufacturers.add(manufacturer)

    # open connection to CDB
    api = CdbApiFactory("http://localhost:8080/cdb")
    try:
        api.authenticateUser("cdb", "cdb")
        api.testAuthenticated()
    except ApiException:
        sys.exit("CDB login failed")
    source_api = api.getSourceApi()

    # create output spreadsheet
    output_book = xlsxwriter.Workbook(args.outputFile)
    output_sheet = output_book.add_worksheet()

    # write output spreadsheet header row
    row_count = 0
    output_sheet.write(row_count, 0, "Name")
    output_sheet.write(row_count, 1, "Description")
    output_sheet.write(row_count, 2, "Contact Info")
    output_sheet.write(row_count, 3, "URL")

    # process each unique manufacturer
    row_count = row_count + 1
    for manufacturer in sorted(manufacturers):

        # check to see if manufacturer exists as a CDB Source
        try:
            mfr_source = source_api.get_source_by_name(manufacturer)
        except ApiException as ex:
            if "ObjectNotFound" not in ex.body:
                print("exception retrieving source for manufacturer: %s - %s" % (manufacturer, ex.body))
            mfr_source = None
        if mfr_source:
            print("source: %s already exists for manufacturer: %s, skipping" % (mfr_source, manufacturer))
        else:
            # write row to output spreadsheet for this manufacturer
            print("adding manufacturer: %s to output spreadsheet" % manufacturer)
            output_sheet.write(row_count, 0, manufacturer)
            row_count = row_count + 1

    # save output spreadsheet
    output_book.close()

    # close CDB connection
    try:
        api.logOutUser()
    except ApiException:
        sys.exit("CDB logout failed")


if __name__ == '__main__':
    main()