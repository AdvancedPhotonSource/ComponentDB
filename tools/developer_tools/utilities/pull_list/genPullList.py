import argparse
import configparser
import gc
import logging
import os
import re
import sys
import warnings
from abc import ABC, abstractmethod
from enum import Enum, auto, unique

import openpyxl
from openpyxl import Workbook

import pandas

CONFIG_SECTION_DEFAULT = "DEFAULT"
CONFIG_RES_DEFAULT_INPUT_DIR = "inputDir"
CONFIG_RES_DEFAULT_OUTPUT_DIR = "outputDir"

CONFIG_SECTION_LOADER = "LOADER"
CONFIG_RES_LOADER_INPUT_FILE = "inputFile"

CONFIG_SECTION_TAGGER = "TAGGING_MODULE"
CONFIG_RES_TAGGER_IGNORE_ERRORS = "ignoreErrors"

CONFIG_SECTION_CALCULATOR = "CALCULATOR_MODULE"
CONFIG_RES_CALCULATOR_PENETRATIONS_FILE = "penetrationsFile"
CONFIG_SECTION_CALCULATOR_SUFFIX = "_CALCULATOR"
CONFIG_RES_CALCULATOR_INPUT_FILE = "inputFile"

CONFIG_SECTION_GENERATOR = "GENERATOR"
CONFIG_RES_GENERATOR_OUTPUT_FILE = "outputFile"

MAGNET_REGEX_STR = "S(\d{2})([AB]):(M|Q|S|FC|FH|FV|SQ|H|V)(\d{1})(T?)"
MAGNET_REGEX = re.compile(MAGNET_REGEX_STR + '$')
MAGNET_TC_REGEX = re.compile(MAGNET_REGEX_STR + ":TC(\d{1})" + '$')
MAGNET_KLIXON_REGEX = re.compile(MAGNET_REGEX_STR + ":TS(\d{1})" + '$')

BIPOLAR_MAGNET_POWER_CABLE_TYPE = "#14/2c (corrector)"

SUFFIX_UNIPOLAR_MAGNET_POWER = "UPOW"
SUFFIX_BIPOLAR_MAGNET_POWER = "BPOW"
SUFFIX_MAGNET_THERMOCOUPLE = "MAGTC"
SUFFIX_KLIXON = "KLIX"

CALCULATOR_SHEET_METHOD = "METHOD"
CALCULATOR_SHEET_PENETRATIONS = "PENETRATIONS"
CALCULATOR_SHEET_LENGTH_TYPE = "LENGTH-TYPE"
CALCULATOR_SHEET_BASE_LENGTHS = "BASE-LENGTHS"
CALCULATOR_SHEET_FROM_LENGTH = "FROM-LENGTH"
CALCULATOR_SHEET_TO_LENGTH = "TO-LENGTH"
CALCULATOR_SHEET_ROUTE = "ROUTE"


@unique
class Field(Enum):
    ROW_VALID = auto()
    ROW_VALID_INFO = auto()
    CDB_CABLE_NAME = auto()
    CDB_CABLE_TECH_SYSTEM = auto()
    CDB_CABLE_DESC = auto()
    CDB_CABLE_ID = auto()
    CDB_CABLE_ENDPOINTS = auto()
    CDB_CABLE_TYPE = auto()
    CDB_CABLE_END1_DEVICE = auto()
    CDB_CABLE_END1_PORT = auto()
    CDB_CABLE_END1_CONNECTOR = auto()
    CDB_CABLE_END2_DEVICE = auto()
    CDB_CABLE_END2_PORT = auto()
    CDB_CABLE_END2_CONNECTOR = auto()
    CDB_CABLE_EXT_NAME = auto()
    CDB_CABLE_IMPORT_ID = auto()
    CDB_CABLE_ALT_ID = auto()
    CDB_CABLE_END1_DESC = auto()
    CDB_CABLE_END2_DESC = auto()
    CDB_CABLE_TYPE_DESC = auto()
    CABLE_ROUTE = auto()
    CABLE_NOTES = auto()
    SRC_LOCATION = auto()
    SRC_DRAWING = auto()
    SRC_END_LENGTH = auto()
    SRC_NOTES = auto()
    DEST_LOCATION = auto()
    DEST_DRAWING = auto()
    DEST_END_LENGTH = auto()
    DEST_NOTES = auto()
    ROUTING_BATCH_ID = auto()
    ROUTING_CABLE_ID = auto()
    ROUTING_CAD_LENGTH = auto()
    ROUTING_ROUTED_LENGTH = auto()
    ROUTING_SECTOR = auto()
    ROUTING_SECTION = auto()
    ROUTING_MAGNET_TYPE = auto()
    ROUTING_MAGNET_NUMBER = auto()
    ROUTING_TRIM = auto()
    ROUTING_TC_NUMBER = auto()
    ROUTING_KLIX_NUMBER = auto()


LABEL_ROW_VALID = "row valid"
LABEL_ROW_VALID_INFO = "row valid info"
LABEL_CDB_CABLE_NAME = "cdb cable name"
LABEL_CDB_CABLE_TECH_SYSTEM = "cdb cable tech system"
LABEL_CDB_CABLE_DESC = "cdb cable description"
LABEL_CDB_CABLE_ID = "cdb cable id"
LABEL_CDB_CABLE_ENDPOINTS = "cdb cable endpoints"
LABEL_CDB_CABLE_TYPE = "cdb cable type"
LABEL_CDB_CABLE_END1_DEVICE = "src device"
LABEL_CDB_CABLE_END1_PORT = "src device port"
LABEL_CDB_CABLE_END1_CONNECTOR = "cdb cable end1 connector"
LABEL_CDB_CABLE_END2_DEVICE = "dest device"
LABEL_CDB_CABLE_END2_PORT = "dest device port"
LABEL_CDB_CABLE_END2_CONNECTOR = "cdb cable end2 connector"
LABEL_CDB_CABLE_EXT_NAME = "cdb cable kabel name"
LABEL_CDB_CABLE_IMPORT_ID = "cdb cable import id"
LABEL_CDB_CABLE_ALT_ID = "cdb cable legacy id"
LABEL_CDB_CABLE_END1_DESC = "cdb cable end1 description"
LABEL_CDB_CABLE_END2_DESC = "cdb cable end2 description"
LABEL_CDB_CABLE_TYPE_DESC = "cdb cable type description"
LABEL_CABLE_ROUTE = "cable route"
LABEL_CABLE_NOTES = "cable notes"
LABEL_SRC_LOCATION = "src location"
LABEL_SRC_DRAWING = "src drawing"
LABEL_SRC_END_LENGTH = "src end length"
LABEL_SRC_NOTES = "src notes"
LABEL_DEST_LOCATION = "dest location"
LABEL_DEST_DRAWING = "dest drawing"
LABEL_DEST_END_LENGTH = "dest end length"
LABEL_DEST_NOTES = "dest notes"
LABEL_ROUTING_CAD_LENGTH = "CAD length"
LABEL_ROUTING_ROUTED_LENGTH = "routed length"


class FieldInfo:

    def __init__(self, key, label):
        self.key = key
        self.label = label

    def get_label(self):
        return self.label


field_info_map = {
    Field.ROW_VALID: FieldInfo(Field.ROW_VALID, LABEL_ROW_VALID),
    Field.ROW_VALID_INFO: FieldInfo(Field.ROW_VALID_INFO, LABEL_ROW_VALID_INFO),
    Field.CDB_CABLE_NAME: FieldInfo(Field.CDB_CABLE_NAME, LABEL_CDB_CABLE_NAME),
    Field.CDB_CABLE_TECH_SYSTEM: FieldInfo(Field.CDB_CABLE_TECH_SYSTEM, LABEL_CDB_CABLE_TECH_SYSTEM),
    Field.CDB_CABLE_DESC: FieldInfo(Field.CDB_CABLE_DESC, LABEL_CDB_CABLE_DESC),
    Field.CDB_CABLE_ID: FieldInfo(Field.CDB_CABLE_ID, LABEL_CDB_CABLE_ID),
    Field.CDB_CABLE_ENDPOINTS: FieldInfo(Field.CDB_CABLE_ENDPOINTS, LABEL_CDB_CABLE_ENDPOINTS),
    Field.CDB_CABLE_TYPE: FieldInfo(Field.CDB_CABLE_TYPE, LABEL_CDB_CABLE_TYPE),
    Field.CDB_CABLE_END1_DEVICE: FieldInfo(Field.CDB_CABLE_END1_DEVICE, LABEL_CDB_CABLE_END1_DEVICE),
    Field.CDB_CABLE_END1_PORT: FieldInfo(Field.CDB_CABLE_END1_PORT, LABEL_CDB_CABLE_END1_PORT),
    Field.CDB_CABLE_END1_CONNECTOR: FieldInfo(Field.CDB_CABLE_END1_CONNECTOR, LABEL_CDB_CABLE_END1_CONNECTOR),
    Field.CDB_CABLE_END2_DEVICE: FieldInfo(Field.CDB_CABLE_END2_DEVICE, LABEL_CDB_CABLE_END2_DEVICE),
    Field.CDB_CABLE_END2_PORT: FieldInfo(Field.CDB_CABLE_END2_PORT, LABEL_CDB_CABLE_END2_PORT),
    Field.CDB_CABLE_END2_CONNECTOR: FieldInfo(Field.CDB_CABLE_END2_CONNECTOR, LABEL_CDB_CABLE_END2_CONNECTOR),
    Field.CDB_CABLE_EXT_NAME: FieldInfo(Field.CDB_CABLE_EXT_NAME, LABEL_CDB_CABLE_EXT_NAME),
    Field.CDB_CABLE_IMPORT_ID: FieldInfo(Field.CDB_CABLE_IMPORT_ID, LABEL_CDB_CABLE_IMPORT_ID),
    Field.CDB_CABLE_ALT_ID: FieldInfo(Field.CDB_CABLE_ALT_ID, LABEL_CDB_CABLE_ALT_ID),
    Field.CDB_CABLE_TYPE_DESC: FieldInfo(Field.CDB_CABLE_TYPE_DESC, LABEL_CDB_CABLE_TYPE_DESC),
    Field.CABLE_ROUTE: FieldInfo(Field.CABLE_ROUTE, LABEL_CABLE_ROUTE),
    Field.CABLE_NOTES: FieldInfo(Field.CABLE_NOTES, LABEL_CABLE_NOTES),
    Field.SRC_LOCATION: FieldInfo(Field.SRC_LOCATION, LABEL_SRC_LOCATION),
    Field.SRC_DRAWING: FieldInfo(Field.SRC_DRAWING, LABEL_SRC_DRAWING),
    Field.SRC_END_LENGTH: FieldInfo(Field.SRC_END_LENGTH, LABEL_SRC_END_LENGTH),
    Field.SRC_NOTES: FieldInfo(Field.SRC_NOTES, LABEL_SRC_NOTES),
    Field.DEST_LOCATION: FieldInfo(Field.DEST_LOCATION, LABEL_DEST_LOCATION),
    Field.DEST_DRAWING: FieldInfo(Field.DEST_DRAWING, LABEL_DEST_DRAWING),
    Field.DEST_END_LENGTH: FieldInfo(Field.DEST_END_LENGTH, LABEL_DEST_END_LENGTH),
    Field.DEST_NOTES: FieldInfo(Field.DEST_NOTES, LABEL_DEST_NOTES),
    Field.ROUTING_CAD_LENGTH: FieldInfo(Field.ROUTING_CAD_LENGTH, LABEL_ROUTING_CAD_LENGTH),
    Field.ROUTING_ROUTED_LENGTH: FieldInfo(Field.ROUTING_ROUTED_LENGTH, LABEL_ROUTING_ROUTED_LENGTH),
}


class ExcelSheetInputModel(ABC):

    def __init__(self, name):
        self.sheet_name = name
        self.sheet = None
        self.workbook = None

    def set_sheet_name(self, name):
        self.sheet_name = name

    def get_sheet_name(self):
        return self.sheet_name

    def set_sheet(self, sheet):
        self.sheet = sheet

    def set_workbook(self, workbook):
        self.workbook = workbook

    def get_workbook_filename(self):
        return self.workbook.get_filename()

    def initialize(self):
        return True, ""


class ExcelRowDictColumnInputModel:

    def __init__(self, key, required=False):
        self.key = key
        self.required = required


class ExcelRowDictSheetInputModel(ExcelSheetInputModel):

    def __init__(self, name, column_specs):
        super().__init__(name)
        self.column_specs = column_specs

    def validate_dimensions(self):

        # log actual dimensions
        num_actual_cols = self.sheet.max_column
        num_actual_rows = self.sheet.max_row
        logging.info("input spreadsheet dimensions: %d cols x %d rows" % (num_actual_cols, num_actual_rows))

        # validate num actual columns matches expected
        num_expected_cols = len(self.column_specs)
        if num_expected_cols != num_actual_cols:
            return False, "sheet: %s actual number of columns %d doesn't match expected number %d" % \
                   (self.sheet_name, num_actual_cols, num_expected_cols)
        else:
            return True, ""

    def initialize(self):
        sheet_init_valid, sheet_init_valid_info = self.validate_dimensions()
        if not sheet_init_valid:
            fatal_error(sheet_init_valid_info)
        return super().initialize()

    def load_data(self):

        print()
        print("Loading file: %s sheet: %s ====================" % (self.get_workbook_filename(), self.sheet_name))
        print()

        records = []
        load_valid = True
        load_valid_info = ""
        first_row_ind = 2
        last_row_ind = self.sheet.max_row

        # get column spec keys
        keys = [spec.key for spec in self.column_specs]
        cols_required = [spec.required for spec in self.column_specs]

        # iterate sheet rows
        rows = self.sheet.rows
        print("loading %d rows" % (last_row_ind+1))
        first_row = True
        row_count = 0
        for row in rows:
            row_count = row_count + 1
            record = {}
            row_valid = True
            row_valid_info = ""
            if first_row:
                # skip first row
                first_row = False
                continue
            for key, cell, required in zip(keys, row, cols_required):
                value = cell.value
                if value is None:
                    value = ""
                if required and value == "":
                    row_valid = False
                    row_valid_info = row_valid_info + "Row: %d missing value for required column %s. " % (row_count, key)
                record[key] = cell.value
            record[Field.ROW_VALID] = row_valid
            record[Field.ROW_VALID_INFO] = row_valid_info
            records.append(record)
            logging.debug("row: %d dict: %s" % (row_count, str(record)))
            if not row_valid:
                load_valid = False
                load_valid_info = "Sheet contains invalid rows."
            if row_count % 1000 == 0:
                print("loaded %d rows" % row_count)

        return load_valid, load_valid_info, records


class ExcelDataFrameSheetInputModel(ExcelSheetInputModel):

    def __init__(self, name):
        super().__init__(name)
        self.frame = None

    def initialize(self):

        (super_init_valid, super_valid_info) = super().initialize()
        if not super_init_valid:
            return False, "base class initialization failed: %s" % super_valid_info

        (load_valid, load_valid_info) = self.load_data()
        if not load_valid:
            return False, "failed to load data for sheet, info: %s" % load_valid_info

        return True, ""

    def load_data(self):

        print()
        print("Loading file: %s sheet: %s ====================" % (self.get_workbook_filename(), self.sheet_name))
        print()

        is_valid = True
        valid_info = ""

        # iterate sheet rows
        rows = self.sheet.rows
        last_row_ind = self.sheet.max_row
        print("loading %d rows" % (last_row_ind + 1))
        row_count = 0
        column_labels = []
        row_labels = []
        row_data = []
        column_labels_next_row = True
        for row in rows:
            row_count = row_count + 1

            # skip blank rows
            blank = True
            for cell in row:
                if cell.value is not None:
                    blank = False
            if blank:
                continue

            # skip comment rows
            cell = row[0]
            if cell.value is not None:
                if cell.value[0] == "#":
                    continue

            # first row after blanks and comments is column labels
            if column_labels_next_row:
                column_labels = [cell.value for cell in row[1:]]
                print("column labels: %s" % str(column_labels))
                column_labels_next_row = False
                continue

            # otherwise row contains row label in first cell and data value in subsequent cells
            else:
                # add row label to list
                row_labels.append(row[0].value)
                # add data values to list of data rows
                row_data.append(cell.value for cell in row[1:])

        # check that row and column labels don't include blanks
        for label in column_labels:
            if label is None or label == "":
                is_valid = False
                valid_info = "blank value in list of column labels"

        for label in row_labels:
            if label is None or label == "":
                is_valid = False
                valid_info = "blank value in list of row labels"

        # create pandas DataFrame with sheet contents
        print("row labels: %s" % str(row_labels))
        self.frame = pandas.DataFrame(row_data, columns=column_labels, index=row_labels)

        return is_valid, valid_info

    def get_data(self):
        return self.frame


class ExcelWorkbookInputModel:

    def __init__(self):
        self.filename = None
        self.sheets = None
        self.workbook = None

    def get_filename(self):
        return self.filename

    def initialize_for_read(self, filename, sheets):

        self.filename = filename
        self.sheets = sheets

        init_valid = True
        init_valid_info = ""

        # create  openpyxl workbook
        with warnings.catch_warnings(record=True):
            warnings.simplefilter("always")
            self.workbook = openpyxl.load_workbook(self.filename, read_only=True, data_only=True)

        # process sheets
        sheet_names = self.workbook.sheetnames
        sheet_index = 0
        for sheet in self.sheets:
            sheet_name = sheet_names[sheet_index]
            if sheet.get_sheet_name() is None:
                sheet.set_sheet_name(sheet_name)
            else:
                if sheet.get_sheet_name() != sheet_name:
                    return False, "sheet name mismatch, expected: %s actual: %s" % (sheet.get_sheet_name(), sheet_name)
            sheet.set_sheet(self.workbook[sheet_name])
            sheet.set_workbook(self)
            (sheet_init_valid, sheet_init_info) = sheet.initialize()
            if not sheet_init_valid:
                return False, "initialization failed for sheet: %s info: %s" % (sheet_name, sheet_init_info)
            sheet_index = sheet_index + 1

        return init_valid, init_valid_info


class ExcelRowDictColumnOutputModel:

    def __init__(self, key):
        self.key = key

    def get_key(self):
        return self.key

    def get_label(self):
        return field_info_map[self.key].get_label()


class ExcelRowDictSheetOutputModel:

    def __init__(self, column_specs, sheet_name):
        self.column_specs = column_specs
        self.sheet_name = sheet_name
        self.sheet = None
        self.workbook = None

    def get_sheet_name(self):
        return self.sheet_name

    def set_sheet(self, sheet):
        self.sheet = sheet

    def set_workbook(self, workbook):
        self.workbook = workbook

    def write_data(self, row_dicts):

        is_valid = True
        valid_info = ""

        # generate header row
        row_ind = 1
        col_ind = 1
        for spec in self.column_specs:
            self.sheet.cell(row=row_ind, column=col_ind).value = spec.get_label()
            col_ind = col_ind + 1

        # generate data rows
        row_ind = 2
        for row_dict in row_dicts:
            col_ind = 1
            for spec in self.column_specs:
                if spec.get_key() not in row_dict:
                    value = ""
                else:
                    value = row_dict[spec.get_key()]
                self.sheet.cell(row=row_ind, column=col_ind).value = value
                col_ind = col_ind + 1
            row_ind = row_ind + 1

        self.workbook.save()

        return is_valid, valid_info


class ExcelWorkbookOutputModel:

    def __init__(self):
        self.filename = None
        self.sheets = None
        self.workbook = None

    def initialize_for_write(self, filename, sheets):

        init_valid = True
        init_valid_info = ""

        self.filename = filename
        self.workbook = Workbook()
        self.sheets = sheets
        sheet_index = 0
        for sheet in self.sheets:
            if sheet_index != 0:
                wb_sheet = self.workbook.create_sheet(sheet.get_sheet_name())
            else:
                wb_sheet = self.workbook.active
                wb_sheet.title = sheet.get_sheet_name()
            sheet.set_sheet(wb_sheet)
            sheet.set_workbook(self)
            sheet_index = sheet_index + 1

        return init_valid, init_valid_info

    def save(self):
        self.workbook.save(self.filename)


class CableInfoLoader:

    def __init__(self, input_dir, config):
        self.input_dir = input_dir
        self.config = config
        self.workbook = None
        self.sheet = None

    def initialize(self):

        init_wb_valid = True
        init_wb_valid_info = ""

        # get input filename from config
        input_filename = get_config_resource(self.config, CONFIG_SECTION_LOADER, CONFIG_RES_LOADER_INPUT_FILE, True)
        file_input = self.input_dir + "/" + input_filename
        if not os.path.isfile(file_input):
            fatal_error("'[%s] inputFile' file: %s does not exist in directory: %s, exiting" %
                        (CONFIG_SECTION_LOADER, input_filename, self.input_dir))

        # create column, sheet, workbook models
        specs = [
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_NAME, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_TECH_SYSTEM, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_DESC, required=False),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_ID, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_ENDPOINTS, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_TYPE, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_END1_DEVICE, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_END1_PORT, required=False),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_END1_CONNECTOR, required=False),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_END2_DEVICE, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_END2_PORT, required=False),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_END2_CONNECTOR, required=False),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_EXT_NAME, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_IMPORT_ID, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_ALT_ID, required=False),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_END1_DESC, required=True),
            ExcelRowDictColumnInputModel(key=Field.CDB_CABLE_END2_DESC, required=True),
        ]
        self.sheet = ExcelRowDictSheetInputModel(name=None, column_specs=specs)
        self.workbook = ExcelWorkbookInputModel()
        (init_wb_valid, init_wb_valid_info) = self.workbook.initialize_for_read(file_input, [self.sheet])

        return init_wb_valid, init_wb_valid_info

    def load_records(self):
        (load_valid, load_valid_info, rows) = self.sheet.load_data()
        return load_valid, load_valid_info, rows

    def finalize(self):
        self.sheet = None
        self.workbook = None
        self.config = None
        self.input_dir = None
        return True, ""


class CableInfoModule(ABC):

    def __init__(self):
        pass

    @abstractmethod
    def initialize(self, config):
        return True, ""

    @abstractmethod
    def process_records(self, records):
        return True, ""

    @abstractmethod
    def finalize(self):
        pass


class TaggingModuleHandler(ABC):

    def __init__(self):
        super().__init__()
        self.handled_record_count = 0

    @abstractmethod
    def get_batch_id(self):
        pass

    @abstractmethod
    # returns True if record is successfully handled.
    def handle(self, record):
        pass

    def increment_handled_record_count(self):
        self.handled_record_count = self.handled_record_count + 1

    def get_handled_record_count(self):
        return self.handled_record_count


class MagnetPowerCableHandler(TaggingModuleHandler, ABC):

    def __init__(self):
        super().__init__()

    def match_regex(self, to_device):
        # match regex
        tagging_dict = {}
        regex_match = MAGNET_REGEX.match(to_device)
        if regex_match:
            sector, section, magnet_type, magnet_number, trim = regex_match.groups()
            cable_id = 'S' + sector + section + '-' + magnet_type + magnet_number + trim + '-' + self.get_batch_id()
            tagging_dict[Field.ROUTING_SECTOR] = sector
            tagging_dict[Field.ROUTING_SECTION] = section
            tagging_dict[Field.ROUTING_MAGNET_TYPE] = magnet_type
            tagging_dict[Field.ROUTING_MAGNET_NUMBER] = magnet_number
            tagging_dict[Field.ROUTING_TRIM] = trim
            return True, self.get_batch_id(), cable_id, tagging_dict
        else:
            return False, None, None, tagging_dict


class UnipolarMagnetPowerCableHandler(MagnetPowerCableHandler):

    def __init__(self):
        super().__init__()
        self.cable_types = {"DLO 444 (pair)", "DLO 535 (pair)", "DLO #2 (pair)", "DLO 4/0 (pair)"}

    def get_batch_id(self):
        return SUFFIX_UNIPOLAR_MAGNET_POWER

    def handle(self, record):

        cable_type = record.get(Field.CDB_CABLE_TYPE, None)
        to_device = record.get(Field.CDB_CABLE_END2_DEVICE, None)

        # check cable type
        if cable_type not in self.cable_types:
            return False, None, None, {}
        else:
            return self.match_regex(to_device)


class BipolarMagnetPowerCableHandler(MagnetPowerCableHandler):

    def __init__(self):
        super().__init__()

    def get_batch_id(self):
        return SUFFIX_BIPOLAR_MAGNET_POWER

    def handle(self, record):

        cable_type = record.get(Field.CDB_CABLE_TYPE, None)
        to_device = record.get(Field.CDB_CABLE_END2_DEVICE, None)

        # check cable type
        if cable_type != BIPOLAR_MAGNET_POWER_CABLE_TYPE:
            return False, None, None, {}
        else:
            return self.match_regex(to_device)


class MagnetThermocoupleCableHandler(TaggingModuleHandler):

    def __init__(self):
        super().__init__()

    def get_batch_id(self):
        return SUFFIX_MAGNET_THERMOCOUPLE

    def handle(self, record):

        to_device = record.get(Field.CDB_CABLE_END2_DEVICE, None)

        tagging_dict = {}
        regex_match = MAGNET_TC_REGEX.match(to_device)
        if regex_match:
            sector, section, magnet_type, magnet_number, trim, tc_number = regex_match.groups()
            cable_id = 'S' + sector + section + '-' + magnet_type + magnet_number + trim + '_TC' + tc_number + '-' + self.get_batch_id()
            tagging_dict[Field.ROUTING_SECTOR] = sector
            tagging_dict[Field.ROUTING_SECTION] = section
            tagging_dict[Field.ROUTING_MAGNET_TYPE] = magnet_type
            tagging_dict[Field.ROUTING_MAGNET_NUMBER] = magnet_number
            tagging_dict[Field.ROUTING_TRIM] = trim
            tagging_dict[Field.ROUTING_TC_NUMBER] = tc_number
            return True, self.get_batch_id(), cable_id, tagging_dict
        else:
            return False, None, None, tagging_dict


class KlixonCableHandler(TaggingModuleHandler):

    def __init__(self):
        super().__init__()

    def get_batch_id(self):
        return SUFFIX_KLIXON

    def handle(self, record):

        to_device = record.get(Field.CDB_CABLE_END2_DEVICE, None)

        tagging_dict = {}
        regex_match = MAGNET_KLIXON_REGEX.match(to_device)
        if regex_match:
            sector, section, magnet_type, magnet_number, trim, klix_number = regex_match.groups()
            cable_id = 'S' + sector + section + '-' + magnet_type + magnet_number + trim + '_TS' + klix_number + '-' + self.get_batch_id()
            tagging_dict[Field.ROUTING_SECTOR] = sector
            tagging_dict[Field.ROUTING_SECTION] = section
            tagging_dict[Field.ROUTING_MAGNET_TYPE] = magnet_type
            tagging_dict[Field.ROUTING_MAGNET_NUMBER] = magnet_number
            tagging_dict[Field.ROUTING_TRIM] = trim
            tagging_dict[Field.ROUTING_KLIX_NUMBER] = klix_number
            return True, self.get_batch_id(), cable_id, tagging_dict
        else:
            return False, None, None, tagging_dict


class TaggingModule(CableInfoModule):

    def __init__(self):
        super().__init__()
        self.handlers = []
        self.sample_data_dict = {}
        self.handling_error_count = 0
        self.total_record_count = 0
        self.invalid_record_count = 0
        self.unhandled_record_count = 0
        self.unhandled_cable_types = set()
        self.handled_record_count = 0
        self.ignore_tagging_errors = False

    def initialize(self, config):

        print()
        print("Tagging Module Initialization ====================")
        print()

        # get config resource for ignoring unhandled items
        self.ignore_tagging_errors = get_config_resource_bool(config, CONFIG_SECTION_TAGGER, CONFIG_RES_TAGGER_IGNORE_ERRORS, False)

        # initialize list of handlers
        self.handlers.append(UnipolarMagnetPowerCableHandler())
        self.handlers.append(BipolarMagnetPowerCableHandler())
        self.handlers.append(MagnetThermocoupleCableHandler())
        self.handlers.append(KlixonCableHandler())

        return True, ""

    # Assigns a routing_batch_id and routing_cable_id to each cable.
    def process_records(self, records):

        for record in records:
            self.total_record_count = self.total_record_count + 1
            is_valid = record.get(Field.ROW_VALID, False)

            if is_valid:
                valid_info = ""
                handled_record = False

                # check for presence of cable type in record
                cable_type = None
                if Field.CDB_CABLE_TYPE not in record:
                    fatal_error("Tagging module encountered record without '%s' field: %s" % (LABEL_CDB_CABLE_TYPE, record))
                else:
                    cable_type = record[Field.CDB_CABLE_TYPE]

                # check for presence of "to device" in record
                to_device = None
                if Field.CDB_CABLE_END2_DEVICE not in record:
                    fatal_error("Tagging module encountered record without '%s' (to device) field: %s" % (LABEL_CDB_CABLE_END2_DEVICE, record))
                else:
                    to_device = record[Field.CDB_CABLE_END2_DEVICE]

                for handler in self.handlers:
                    handler_name = type(handler).__name__
                    (handled_record, batch_id, cable_id, tagging_dict) = handler.handle(record)

                    if handled_record:

                        self.handled_record_count = self.handled_record_count + 1

                        if batch_id is None or len(batch_id) == 0:
                            is_valid = False
                            valid_info = valid_info + "Handler: %s failed to set batch id. " % handler_name
                        elif cable_id is None or len(cable_id) == 0:
                            is_valid = False
                            valid_info = valid_info + "Handler: %s failed to set cable id" % handler_name

                        if is_valid:
                            handler.increment_handled_record_count()
                            record[Field.ROUTING_BATCH_ID] = batch_id
                            record[Field.ROUTING_CABLE_ID] = cable_id
                            record.update(tagging_dict)
                            if handler_name not in self.sample_data_dict:
                                self.sample_data_dict[handler_name] = (batch_id, cable_id)
                        else:
                            self.handling_error_count = self.handling_error_count + 1
                            record[Field.ROW_VALID] = False
                            record[Field.ROW_VALID_INFO] = valid_info

                        break

                if not handled_record:
                    self.unhandled_record_count = self.unhandled_record_count + 1
                    self.unhandled_cable_types.add(cable_type)
                    record[Field.ROW_VALID] = False
                    record[Field.ROW_VALID_INFO] = "Tagging module failed to handle record."

            else:
                self.invalid_record_count = self.invalid_record_count + 1

        is_valid = True
        valid_info = ""

        if self.unhandled_record_count > 0 or self.handling_error_count > 0:
            if not self.ignore_tagging_errors:
                is_valid = False
                valid_info = "Tagging module unhandled records: %d handling errors: %d" % (self.unhandled_record_count, self.handling_error_count)

        return is_valid, valid_info

    def finalize(self):

        print()
        print("Tagging Module Stats ====================")
        print()
        print("Total records: %d" % self.total_record_count)
        print("Handled records: %d" % self. handled_record_count)
        print("Unhandled records: %d" % self.unhandled_record_count)
        print("Handling errors: %d" % self.handling_error_count)

        for handler in self.handlers:
            handler_name = type(handler).__name__
            handled_record_count = handler.get_handled_record_count()
            print()
            print("Handler: %s (batch id: %s) records handled: %d " % (handler_name, handler.get_batch_id(), handled_record_count))
            if handled_record_count > 0:
                (batch_id, cable_id) = self.sample_data_dict[handler_name]
                print("sample routing cable id: %s" % cable_id)

        self.handlers = None

        if len(self.unhandled_cable_types) > 0:
            print()
            print("Unhandled cable types:")
            print()
            for cable_type in self.unhandled_cable_types:
                print(cable_type)


class LengthCalculator:

    def __init__(self, batch_id, penetrations_dataframe):
        super().__init__()
        self.batch_id = batch_id
        self.config = None
        self.sheet_names = None
        self.workbook_data_dict = {}
        self.penetrations_dataframe = penetrations_dataframe

    def get_config_section(self):
        if self.batch_id is None:
            return ""
        else:
            return self.batch_id + CONFIG_SECTION_CALCULATOR_SUFFIX

    def initialize(self, config):

        is_valid = True
        valid_info = ""
        self.config = config

        config_resource = CONFIG_RES_CALCULATOR_INPUT_FILE
        config_section = self.get_config_section()
        input_filename = get_config_resource(self.config, config_section, config_resource, True)
        if input_filename is None:
            is_valid = False
            valid_info = "failed to get input file config resource: %s from section: %s" % (config_resource, config_section)
        input_file = get_option_input_dir() + "/" + input_filename

        # initialize and read workbook
        self.sheet_names = [CALCULATOR_SHEET_METHOD, CALCULATOR_SHEET_PENETRATIONS, CALCULATOR_SHEET_LENGTH_TYPE, CALCULATOR_SHEET_BASE_LENGTHS, CALCULATOR_SHEET_FROM_LENGTH, CALCULATOR_SHEET_TO_LENGTH, CALCULATOR_SHEET_ROUTE]
        sheet_list = []
        for sheet_name in self.sheet_names:
            sheet = ExcelDataFrameSheetInputModel(sheet_name)
            sheet_list.append(sheet)
        workbook = ExcelWorkbookInputModel()
        (init_wb_valid, init_wb_valid_info) = workbook.initialize_for_read(input_file, sheet_list)
        if not init_wb_valid:
            is_valid = False
            valid_info = "failed to initialize workbook for %s calculator, details: %s" % (self.batch_id, init_wb_valid_info)

        # create dict with workbook sheet data in pandas data frames
        for sheet in sheet_list:
            self.workbook_data_dict[sheet.get_sheet_name()] = sheet.get_data()

        return is_valid, valid_info

    def handle_record(self, record):

        is_valid = True
        valid_info = ""

        # get tagging field values for lookups
        
        if Field.ROUTING_SECTOR not in record or record[Field.ROUTING_SECTOR] == "":
            is_valid = False
            valid_info = "record is missing sector number"
            # return failure code or just mark record invalid?  for now, return failure code
            return is_valid, valid_info
        else:
            sector_id = "S" + record[Field.ROUTING_SECTOR]
        
        if Field.ROUTING_SECTION not in record or record[Field.ROUTING_SECTION] == "":
            is_valid = False
            valid_info = "record is missing section"
            return is_valid, valid_info
        else:
            section_id = record[Field.ROUTING_SECTION]
            
        if Field.ROUTING_MAGNET_TYPE not in record or record[Field.ROUTING_MAGNET_TYPE] == "":
            is_valid = False
            valid_info = "record is missing magnet type"
            return is_valid, valid_info
        else:
            magnet_type = record[Field.ROUTING_MAGNET_TYPE]

        if Field.ROUTING_MAGNET_NUMBER not in record or record[Field.ROUTING_MAGNET_NUMBER] == "":
            is_valid = False
            valid_info = "record is missing magnet number"
            return is_valid, valid_info
        else:
            magnet_number = record[Field.ROUTING_MAGNET_NUMBER]

        device_type_id = section_id + magnet_type + magnet_number

        # determine calculation method
        method_sheet = self.workbook_data_dict[CALCULATOR_SHEET_METHOD]
        calculation_method = method_sheet[device_type_id][sector_id]
        if pandas.isnull(calculation_method) or calculation_method == "":
            is_valid = False
            valid_info = "missing value in sheet: %s for device id: %s sector id: %s" % (CALCULATOR_SHEET_METHOD, device_type_id, sector_id)
            return is_valid, valid_info

        if calculation_method == "A":

            # get creo length
            length_type_sheet = self.workbook_data_dict[CALCULATOR_SHEET_LENGTH_TYPE]
            length_type_value = length_type_sheet[device_type_id][sector_id]
            if pandas.isnull(length_type_value) or length_type_value == "":
                is_valid = False
                valid_info = "missing value in sheet: %s for device id: %s sector id: %s" % (CALCULATOR_SHEET_LENGTH_TYPE, device_type_id, sector_id)
                return is_valid, valid_info
            base_lengths_sheet = self.workbook_data_dict[CALCULATOR_SHEET_BASE_LENGTHS]
            creo_length_value = base_lengths_sheet[device_type_id][length_type_value]
            if pandas.isnull(creo_length_value) or creo_length_value == "":
                is_valid = False
                valid_info = "missing value in sheet: %s for device id: %s length type: %s" % (CALCULATOR_SHEET_BASE_LENGTHS, device_type_id, length_type_value)
                return is_valid, valid_info

            # get penetration length
            penetrations_sheet = self.workbook_data_dict[CALCULATOR_SHEET_PENETRATIONS]
            penetration_name_value = penetrations_sheet[device_type_id][sector_id]
            if pandas.isnull(penetration_name_value) or penetration_name_value == "":
                is_valid = False
                valid_info = "missing value in sheet: %s for device id: %s sector id: %s" % (CALCULATOR_SHEET_PENETRATIONS, device_type_id, sector_id)
                return is_valid, valid_info
            penetration_length_value = 0
            if penetration_name_value != "NONE":
                penetration_tokens = penetration_name_value.split('-')
                penetration_sector = penetration_tokens[0]
                penetration_column = penetration_tokens[1]
                penetrations_length_sheet = self.penetrations_dataframe
                penetration_length_value = penetrations_length_sheet[penetration_column][penetration_sector]
                if pandas.isnull(penetration_length_value) or penetration_length_value == "":
                    is_valid = False
                    valid_info = "missing value in penetration lengths file for column: %s sector: %s" % (penetration_column, penetration_sector)
                    return is_valid, valid_info

            # get end lengths
            from_length_sheet = self.workbook_data_dict[CALCULATOR_SHEET_FROM_LENGTH]
            from_end_length_value = from_length_sheet[device_type_id][sector_id]
            if pandas.isnull(from_end_length_value) or from_end_length_value == "":
                is_valid = False
                valid_info = "missing value in sheet: %s for device id: %s sector id: %s" % (CALCULATOR_SHEET_FROM_LENGTH, device_type_id, sector_id)
                return is_valid, valid_info
            to_length_sheet = self.workbook_data_dict[CALCULATOR_SHEET_TO_LENGTH]
            to_end_length_value = to_length_sheet[device_type_id][sector_id]
            if pandas.isnull(to_end_length_value) or to_end_length_value == "":
                is_valid = False
                valid_info = "missing value in sheet: %s for device id: %s sector id: %s" % (CALCULATOR_SHEET_TO_LENGTH, device_type_id, sector_id)
                return is_valid, valid_info

            # calculate length
            calculated_length_value = int(creo_length_value) + int(penetration_length_value) + int(from_end_length_value) + int(to_end_length_value)

            # set calculated length in record
            record[Field.ROUTING_ROUTED_LENGTH] = calculated_length_value

        else:
            is_valid = False
            valid_info = "calculation method not implemented: %s" % calculation_method
            return is_valid, valid_info

        return is_valid, valid_info

    def finalize(self):
        pass


class LengthCalculatorModule(CableInfoModule):

    def __init__(self):
        super().__init__()
        self.config = None
        self.calculator_dict = {}
        self.penetrations_dataframe = None

    def initialize(self, config):

        print()
        print("Calculator Module Initialization ====================")
        print()

        is_valid = True
        valid_info = ""
        self.config = config

        # read penetrations file so it can be shared with all calculators in project
        penetrations_config_resource = CONFIG_RES_CALCULATOR_PENETRATIONS_FILE
        config_section = CONFIG_SECTION_CALCULATOR
        penetrations_filename = get_config_resource(config, config_section, penetrations_config_resource, True)
        if penetrations_filename is None:
            is_valid = False
            valid_info = "failed to get penetrations file config resource: %s from section: %s" % (penetrations_config_resource, config_section)
        penetrations_file = get_option_input_dir() + "/" + penetrations_filename

        # initialize and read workbook
        sheet = ExcelDataFrameSheetInputModel(None)
        workbook = ExcelWorkbookInputModel()
        (init_wb_valid, init_wb_valid_info) = workbook.initialize_for_read(penetrations_file, [sheet])
        if not init_wb_valid:
            is_valid = False
            valid_info = "failed to initialize calculator penetrations workbook, details: %s" % init_wb_valid_info

        # create dict with workbook sheet data in pandas data frames
        self.penetrations_dataframe = sheet.get_data()

        # create and initialize calculators
        calculator_names = [SUFFIX_UNIPOLAR_MAGNET_POWER]
        for name in calculator_names:
            calculator = LengthCalculator(name, self.penetrations_dataframe)
            (init_valid, init_valid_info) = calculator.initialize(config)
            if not init_valid:
                is_valid = False
                valid_info = init_valid_info
                break
            else:
                self.calculator_dict[name] = calculator

        return is_valid, valid_info

    def process_records(self, records):

        is_valid = True
        valid_info = ""

        for record in records:

            if Field.ROUTING_BATCH_ID not in record:
                return False, "record missing batch ID"
            else:
                batch_id = record[Field.ROUTING_BATCH_ID]

            if batch_id not in self.calculator_dict:
                return False, "no calculator for batch id: %s" % batch_id

            calculator = self.calculator_dict[batch_id]
            (handle_valid, handle_valid_info) = calculator.handle_record(record)
            if not handle_valid:
                return False, "calculator: %s failed to handle cable: %s reason: %s" % (batch_id, record[Field.CDB_CABLE_NAME], handle_valid_info)

            print(record[Field.ROUTING_ROUTED_LENGTH])

        return is_valid, valid_info

    def finalize(self):
        pass


class PullListGenerator:

    def __init__(self, output_dir, config):
        self.output_dir = output_dir
        self.config = config
        self.workbook = None
        self.sheet = None

    def initialize(self):

        print()
        print("Pull List Geneator Initialization ====================")
        print()

        init_valid = True
        init_valid_info = ""

        # get output filename from config
        output_filename = get_config_resource(self.config, CONFIG_SECTION_GENERATOR, CONFIG_RES_GENERATOR_OUTPUT_FILE, True)
        file_output = self.output_dir + "/" + output_filename

        # create column, sheet, workbook models
        specs = [
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_NAME),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_EXT_NAME),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_TYPE),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_TYPE_DESC),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_TECH_SYSTEM),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_ALT_ID),
            ExcelRowDictColumnOutputModel(key=Field.CABLE_ROUTE),
            ExcelRowDictColumnOutputModel(key=Field.CABLE_NOTES),
            ExcelRowDictColumnOutputModel(key=Field.SRC_LOCATION),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_END1_DEVICE),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_END1_PORT),
            ExcelRowDictColumnOutputModel(key=Field.SRC_DRAWING),
            ExcelRowDictColumnOutputModel(key=Field.SRC_END_LENGTH),
            ExcelRowDictColumnOutputModel(key=Field.SRC_NOTES),
            ExcelRowDictColumnOutputModel(key=Field.DEST_LOCATION),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_END2_DEVICE),
            ExcelRowDictColumnOutputModel(key=Field.CDB_CABLE_END2_PORT),
            ExcelRowDictColumnOutputModel(key=Field.DEST_DRAWING),
            ExcelRowDictColumnOutputModel(key=Field.DEST_END_LENGTH),
            ExcelRowDictColumnOutputModel(key=Field.DEST_NOTES),
            ExcelRowDictColumnOutputModel(key=Field.ROUTING_CAD_LENGTH),
            ExcelRowDictColumnOutputModel(key=Field.ROUTING_ROUTED_LENGTH),
        ]
        self.sheet = ExcelRowDictSheetOutputModel(specs, "pull list")
        self.workbook = ExcelWorkbookOutputModel()
        (init_valid, init_valid_info) = self.workbook.initialize_for_write(file_output, [self.sheet])

        return init_valid, init_valid_info

    def generate(self, records):
        return self.sheet.write_data(records)


def fatal_error(error_msg):
    print()
    print("ERROR ====================")
    print()
    print(error_msg)
    sys.exit(0)


def get_config_resource(config, section, key, is_required, print_value=True, print_mask=None):
    value = None
    if section not in config:
        fatal_error("Invalid config section: %s, exiting" % section)
    if key not in config[section]:
        if is_required:
            fatal_error("Config key: %s not found in section: %s, exiting" % (key, section))
    else:
        value = config[section][key]
        if is_required and len(value) == 0:
            fatal_error("value not provided for required option '[%s] %s', exiting" % (section, key))
    if print_value:
        print_obj = value
        if value is not None:
            if print_mask is not None:
                if len(value) > 0:
                    print_obj = print_mask
        print("[%s] %s: %s" % (section, key, print_obj))
    return value


def get_config_resource_bool(config, section, key, is_required):
    config_value = get_config_resource(config, section, key, is_required)
    if config_value in ("True", "TRUE", "true", "Yes", "YES", "yes", "On", "ON", "on", "1"):
        return True
    else:
        return False
    
    
option_input_dir = None
option_output_dir = None


def get_option_input_dir():
    global option_input_dir
    return option_input_dir


def get_option_output_dir():
    global option_output_dir
    return option_output_dir


def main():

    # parse command line args
    parser = argparse.ArgumentParser()
    parser.add_argument("--configDir", help="Directory containing script config files.", required=True)
    args = parser.parse_args()

    print()
    print("==================== genPullList.py ====================")
    print()
    print("configDir: %s" % args.configDir)

    #
    # Determine config file names and paths and test.
    #

    option_config_dir = args.configDir
    if not os.path.isdir(option_config_dir):
        fatal_error("Specified configDir: %s does not exist, exiting" % option_config_dir)

    file_config_main = option_config_dir + "/genPullList.conf"
    if not os.path.isfile(file_config_main):
        fatal_error("'genPullList.conf' file not found in configDir: %s', exiting" % option_config_dir)

    #
    # Process options.
    #

    # read config files
    config = configparser.ConfigParser()
    config.read(file_config_main)

    print()
    print("preimport.conf OPTIONS ====================")
    print()

    # process inputDir option
    global option_input_dir
    option_input_dir = get_config_resource(config, CONFIG_SECTION_DEFAULT, CONFIG_RES_DEFAULT_INPUT_DIR, True)
    if not os.path.isdir(option_input_dir):
        fatal_error("'[%s] inputDir' directory: %s does not exist, exiting" % ('DEFAULT', option_input_dir))

    # process outputDir option
    global option_output_dir
    option_output_dir = get_config_resource(config, CONFIG_SECTION_DEFAULT, CONFIG_RES_DEFAULT_OUTPUT_DIR, True)
    if not os.path.isdir(option_output_dir):
        fatal_error("'[%s] outputDir' directory: %s does not exist, exiting" % ('DEFAULT', option_output_dir))

    #
    # Configure logging.
    #

    # log file
    file_log = "%s/genPullList.log" % option_output_dir

    # configure logging
    logging.basicConfig(filename=file_log, filemode='w', level=logging.DEBUG, format='%(levelname)s - %(message)s')


    #
    # Load cable records.
    #

    loader = CableInfoLoader(option_input_dir, config)

    (loader_init_valid, loader_init_valid_info) = loader.initialize()
    if not loader_init_valid:
        fatal_error("CableInfoLoader initialization failed: " + loader_init_valid_info)

    (load_valid, load_valid_info, cable_records) = loader.load_records()
    if not load_valid:
        fatal_error("CableInfoLoader loader failed: " + load_valid_info)

    (loader_fini_valid, loader_fini_valid_info) = loader.finalize()
    if not loader_fini_valid:
        fatal_error("CableInfoLoader initialization failed: " + loader_fini_valid_info)

    # run garbage collection
    loader = None
    unreachable = gc.collect()
    logging.debug("gc loader unreachable: " + str(unreachable))
    logging.debug("gc loader stats: " + str(gc.get_stats()))

    #
    # Invoke each CableModule for each cable record.
    #
    module_list = [TaggingModule(), LengthCalculatorModule()]
    for module in module_list:
        (module_init_valid, module_init_valid_info) = module.initialize(config)
        if not module_init_valid:
            fatal_error("Module initialization failed: " + module_init_valid_info)
        (process_valid, process_valid_info) = module.process_records(cable_records)
        if not process_valid:
            fatal_error("Module processing failed: " + process_valid_info)
        module.finalize()
        # run garbage collection after each module
        module = None
        unreachable = gc.collect()
        logging.debug("gc loader unreachable: " + str(unreachable))
        logging.debug("gc loader stats: " + str(gc.get_stats()))

    #
    # run garbage collection after all modules
    #
    module_list = None
    unreachable = gc.collect()
    logging.debug("gc loader unreachable: " + str(unreachable))
    logging.debug("gc loader stats: " + str(gc.get_stats()))

    #
    # Generate pull list output for each cable record.
    #
    generator = PullListGenerator(option_output_dir, config)
    (generator_init_valid, generator_init_valid_info) = generator.initialize()
    if not generator_init_valid:
        fatal_error("Module initialization failed: " + generator_init_valid_info)
    (generate_valid, generate_valid_info) = generator.generate(cable_records)
    if not generate_valid:
        fatal_error("Module processing failed: " + generate_valid_info)


if __name__ == '__main__':
    main()